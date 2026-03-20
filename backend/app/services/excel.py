"""
Excel Processing Service
Stage 1  — validate file (extension, size)
Stage 2  — parse sheets with pandas
Stage 3  — validate rows (types, ranges, duplicates)
Stage 4  — check overwrite
Stage 5  — DB upsert (find/create employee, delete old record if overwrite)
Stage 6  — scoring pipeline (score → trend → confidence → suggestion)
Stage 7  — save upload history + commit
"""
import io
from datetime import date
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.models.employee import Employee
from app.models.performance import MonthlyPerformance, UploadHistory
from app.models.suggestion import Suggestion
from app.schemas.upload import NameChange
from app.services.confidence import calculate_confidence
from app.services.scoring import calculate_score
from app.services.suggestion import build_full_explanation, generate_suggestion
from app.services.trend import analyze_trend
from app.services.weights import get_expected_columns, get_roles

MONTH_LABELS = {
    1: "January",  2: "February", 3: "March",    4: "April",
    5: "May",      6: "June",     7: "July",      8: "August",
    9: "September",10: "October", 11: "November", 12: "December",
}


# ── validation helpers ────────────────────────────────────────────────────────

def _coerce(val: Any, default: float = 0.0) -> float:
    """Safely convert any value to float."""
    if val is None:
        return default
    if isinstance(val, float) and pd.isna(val):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _validate_file(file_bytes: bytes, filename: str, max_mb: int) -> List[str]:
    errors: List[str] = []
    if not filename.lower().endswith(".xlsx"):
        errors.append("Only .xlsx files are accepted.")
        return errors
    if len(file_bytes) > max_mb * 1024 * 1024:
        errors.append(f"File too large (max {max_mb} MB).")
    return errors


def _validate_row(row: Dict, row_num: int) -> Tuple[bool, List[str]]:
    errors: List[str] = []

    emp_id = str(row.get("employee_id", "")).strip()
    name   = str(row.get("name",        "")).strip()

    if not emp_id:
        errors.append(f"Row {row_num}: employee_id is empty.")
    if not name:
        errors.append(f"Row {row_num}: name is empty.")

    try:
        dp = int(_coerce(row.get("days_present", 0)))
        da = int(_coerce(row.get("days_assigned", 1)))
        if not (0 <= dp <= 31):
            errors.append(f"Row {row_num}: days_present must be 0-31 (got {dp}).")
        if not (1 <= da <= 31):
            errors.append(f"Row {row_num}: days_assigned must be 1-31 (got {da}).")
        if dp > da:
            errors.append(
                f"Row {row_num}: days_present ({dp}) cannot exceed days_assigned ({da})."
            )
    except Exception:
        errors.append(f"Row {row_num}: days_present / days_assigned must be numbers.")

    mr = _coerce(row.get("manager_rating"), default=-1.0)
    if not (1.0 <= mr <= 5.0):
        errors.append(f"Row {row_num}: manager_rating must be 1-5 (got {mr}).")

    return len(errors) == 0, errors


# ── main upload pipeline ──────────────────────────────────────────────────────

def process_upload(
    file_bytes:    bytes,
    filename:      str,
    business_id:   int,
    business_type: str,
    month:         int,
    year:          int,
    overwrite:     bool,
    db:            Session,
    max_mb:        int = 5,
) -> Tuple[bool, Dict, List[str]]:
    """
    Returns:
        (True,  summary_dict,           warnings_list)   — success
        (False, {"needs_confirmation"}, [])               — overwrite needed
        (False, {},                     [error_strings])  — hard failure
    """

    # ── Stage 1: file validation ──────────────────────────────────────────────
    file_errors = _validate_file(file_bytes, filename, max_mb)
    if file_errors:
        return False, {}, file_errors

    month_str = f"{year}-{month:02d}"

    # ── Stage 2: parse Excel ──────────────────────────────────────────────────
    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
    except Exception as e:
        return False, {}, [f"Cannot read Excel file: {e}"]

    valid_roles = get_roles(business_type)
    role_sheets = [s for s in xls.sheet_names if s.lower() != "instructions"]

    if not role_sheets:
        sample = ", ".join(
            r.replace("_", " ").title() for r in valid_roles[:3]
        )
        return False, {}, [
            f"No role sheets found. Sheets must be named by role "
            f"(e.g. {sample})."
        ]

    # ── Stage 3: validate all rows ────────────────────────────────────────────
    upload_rows:   List[Dict]      = []
    global_errors: List[str]       = []
    warnings:      List[str]       = []
    seen_ids:      Dict[str, int]  = {}

    for sheet_name in role_sheets:
        role_key = sheet_name.strip().lower().replace(" ", "_")

        if role_key not in valid_roles:
            warnings.append(
                f"Sheet '{sheet_name}' does not match any role for "
                f"'{business_type}'. Skipping. "
                f"Valid roles: {', '.join(valid_roles)}."
            )
            continue

        try:
            df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
        except Exception as e:
            global_errors.append(f"Sheet '{sheet_name}': cannot read — {e}")
            continue

        # normalise column names
        df.columns = [
            str(c).strip().lower().replace(" ", "_") for c in df.columns
        ]

        if df.empty:
            continue

        for idx, raw_row in df.iterrows():
            row_dict: Dict[str, Any] = {
                str(k): v for k, v in raw_row.to_dict().items()
            }
            row_num = int(str(idx)) + 2   # Excel row number (1-based + header)

            ok, errs = _validate_row(row_dict, row_num)
            if not ok:
                global_errors.extend(errs)
                continue

            emp_id = str(row_dict.get("employee_id", "")).strip().upper()

            # duplicate employee_id check across entire file
            if emp_id in seen_ids:
                global_errors.append(
                    f"Duplicate employee_id '{emp_id}' — "
                    f"sheet '{sheet_name}' row {row_num} "
                    f"(first seen row {seen_ids[emp_id]})."
                )
                continue
            seen_ids[emp_id] = row_num

            # build clean metrics dict
            dp  = int(_coerce(row_dict.get("days_present",  0)))
            da  = int(_coerce(row_dict.get("days_assigned", 1)))
            mr  = _coerce(row_dict.get("manager_rating", 3.0))
            att = round((dp / da) * 100, 2) if da > 0 else 0.0

            metrics: Dict[str, Any] = {
                "days_present":   dp,
                "days_assigned":  da,
                "manager_rating": mr,
                "attendance_pct": att,
            }

            expected_cols = get_expected_columns(business_type, role_key)
            skip_cols = {
                "employee_id", "name", "role",
                "days_present", "days_assigned", "manager_rating",
            }
            for col in expected_cols:
                if col in skip_cols:
                    continue
                raw_val = row_dict.get(col)
                if raw_val is None:
                    metrics[col] = None
                elif isinstance(raw_val, float) and pd.isna(raw_val):
                    metrics[col] = None
                else:
                    try:
                        metrics[col] = float(str(raw_val).strip())
                    except ValueError:
                        metrics[col] = None

            upload_rows.append({
                "employee_id": emp_id,
                "name":        str(row_dict.get("name", "")).strip(),
                "role":        role_key,
                "metrics":     metrics,
            })

    if global_errors:
        return False, {}, global_errors

    if not upload_rows:
        return False, {}, ["No valid employee rows found in the file."]

    # ── Stage 4: overwrite check ──────────────────────────────────────────────
    existing_count = (
        db.query(MonthlyPerformance)
        .join(Employee)
        .filter(
            Employee.business_id == business_id,
            MonthlyPerformance.month == month_str,
        )
        .count()
    )
    if existing_count > 0 and not overwrite:
        return False, {
            "needs_confirmation": True,
            "month_str":          month_str,
            "existing_count":     existing_count,
        }, []

    # ── Stage 5 + 6: DB upsert + scoring pipeline ────────────────────────────
    summary: Dict[str, Any] = {
        "total_processed":   0,
        "new_employees":     [],
        "absent_employees":  [],
        "two_month_missing": [],
        "name_changes":      [],
        "reactivations":     [],
    }

    upload_ids: set = {r["employee_id"] for r in upload_rows}

    # pre-load all employees for this business to avoid N+1 queries
    active_employees = db.query(Employee).filter(
        Employee.business_id == business_id,
        Employee.status == "active",
    ).all()
    inactive_employees = db.query(Employee).filter(
        Employee.business_id == business_id,
        Employee.status == "inactive",
    ).all()

    active_map:   Dict[str, Any] = {e.employee_id: e for e in active_employees}
    inactive_map: Dict[str, Any] = {e.employee_id: e for e in inactive_employees}

    for row in upload_rows:
        emp_id:  str            = row["employee_id"]
        name:    str            = row["name"]
        role:    str            = row["role"]
        metrics: Dict[str, Any] = row["metrics"]

        # ── find or create employee ───────────────────────────────────────────
        emp: Optional[Employee] = None

        if emp_id in active_map:
            emp = active_map[emp_id]
            if emp.name != name:
                summary["name_changes"].append(
                    NameChange(
                        employee_id=emp_id,
                        old_name=str(emp.name),
                        new_name=name,
                    )
                )
                emp.name = name                       # type: ignore[assignment]
            emp.role                = role            # type: ignore[assignment]
            emp.consecutive_missing = 0               # type: ignore[assignment]

        elif emp_id in inactive_map:
            emp                     = inactive_map[emp_id]
            emp.status              = "active"        # type: ignore[assignment]
            emp.date_left           = None            # type: ignore[assignment]
            emp.consecutive_missing = 0               # type: ignore[assignment]
            emp.name                = name            # type: ignore[assignment]
            emp.role                = role            # type: ignore[assignment]
            summary["reactivations"].append(emp_id)
            active_map[emp_id]      = emp

        else:
            emp = Employee(
                business_id=business_id,
                employee_id=emp_id,
                name=name,
                role=role,
                status="active",
                consecutive_missing=0,
                date_joined=date.today(),
            )
            db.add(emp)
            db.flush()   # get emp.id before creating performance record
            summary["new_employees"].append(emp_id)
            active_map[emp_id] = emp

        # ── delete existing records if overwrite ──────────────────────────────
        if overwrite:
            db.query(MonthlyPerformance).filter(
                MonthlyPerformance.employee_id == emp.id,
                MonthlyPerformance.month == month_str,
            ).delete(synchronize_session=False)

            db.query(Suggestion).filter(
                Suggestion.employee_id == emp.id,
                Suggestion.month == month_str,
            ).delete(synchronize_session=False)

            db.flush()

        # ── scoring ───────────────────────────────────────────────────────────
        overall, breakdown, zero_viol, red_flags = calculate_score(
            metrics, business_type, role
        )

        perf = MonthlyPerformance(
            employee_id=emp.id,
            month=month_str,
            year=year,
            role=role,
            metrics=metrics,
            overall_score=overall,
            metric_breakdown=breakdown,
        )
        db.add(perf)
        db.flush()

        # ── trend ─────────────────────────────────────────────────────────────
        history_records = (
            db.query(MonthlyPerformance)
            .filter(MonthlyPerformance.employee_id == emp.id)
            .order_by(MonthlyPerformance.month)
            .all()
        )
        history = [
            {"month": r.month, "score": r.overall_score or 0}
            for r in history_records
        ]
        trend_data     = analyze_trend(history)
        months_tracked = len(history)

        # ── confidence ────────────────────────────────────────────────────────
        expected_cols   = get_expected_columns(business_type, role)
        confidence_data = calculate_confidence(
            months_tracked=months_tracked,
            metrics=metrics,
            expected_columns=expected_cols,
            volatility=trend_data.get("volatility"),
            red_flags=red_flags,
        )

        # ── suggestion ────────────────────────────────────────────────────────
        sug_result = generate_suggestion(
            overall_score=overall,
            trend=trend_data["trend"],
            red_flags=red_flags,
            zero_tolerance=zero_viol,
            months_tracked=months_tracked,
        )

        explanation = build_full_explanation(
            metrics=metrics,
            overall_score=overall,
            breakdown=breakdown,
            trend_data=trend_data,
            red_flags=red_flags,
            confidence_data=confidence_data,
            suggestion_result=sug_result,
            months_tracked=months_tracked,
        )

        sug = Suggestion(
            employee_id=emp.id,
            month=month_str,
            year=year,
            suggestion=sug_result["suggestion"],
            confidence_score=confidence_data["confidence_score"],
            trend=trend_data["trend"],
            explanation=explanation,
            red_flags=red_flags if red_flags else None,
            auto_triggered=sug_result.get("auto_triggered"),
        )
        db.add(sug)
        summary["total_processed"] += 1

    # ── absent employees (active in DB but not in this upload) ────────────────
    for emp_id, emp in active_map.items():
        if emp_id not in upload_ids:
            current_missing = int(emp.consecutive_missing or 0)
            emp.consecutive_missing = current_missing + 1   # type: ignore[assignment]
            summary["absent_employees"].append(emp_id)
            if (current_missing + 1) >= 2:
                summary["two_month_missing"].append(emp_id)

    # ── Stage 7: upload history + commit ──────────────────────────────────────
    history_rec = UploadHistory(
        business_id=business_id,
        month=month_str,
        year=year,
        filename=filename,
        employees_processed=summary["total_processed"],
        new_employees=len(summary["new_employees"]),
        absent_employees=len(summary["absent_employees"]),
        status="success",
    )
    db.add(history_rec)
    db.commit()

    return True, summary, warnings


# ── template generator ────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SAMPLE_FILL = PatternFill("solid", fgColor="E8F4FD")
_INST_FILL   = PatternFill("solid", fgColor="FFF3CD")


def _sample_val(col: str, role: str, sheet_index: int = 0) -> Any:
    """Return a sensible sample value for each column."""
    # unique EMP ID per role sheet — EMP001, EMP002, EMP003...
    sample_emp_id = f"EMP{(sheet_index + 1):03d}"

    mapping: Dict[str, Any] = {
        "employee_id":    sample_emp_id,
        "name":           f"Sample Employee {sheet_index + 1}",
        "role":           role,
        "days_present":   25,
        "days_assigned":  26,
        "manager_rating": 4.0,
    }
    if col in mapping:
        return mapping[col]
    if col.endswith("_pct") or col.endswith("_score"):
        return 85
    if any(w in col for w in (
        "error", "violation", "complaint",
        "incident", "failure", "missing",
    )):
        return 0
    if "target" in col:
        return 100000
    if "achieved" in col:
        return 90000
    return 0


def generate_template(
    business_type: str,
    business_name: str,
    month:         int,
    year:          int,
) -> bytes:
    """Returns raw bytes of a styled .xlsx template for the given business type."""
    wb = openpyxl.Workbook()

    # remove default sheet safely
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    # ── Instructions sheet ────────────────────────────────────────────────────
    wi = wb.create_sheet("Instructions")
    wi["A1"] = "Decision Intelligence Framework — Upload Template"
    wi["A1"].font = Font(bold=True, size=14)
    wi["A3"] = f"Business: {business_name}"
    wi["A4"] = f"Month: {MONTH_LABELS.get(month, str(month))} {year}"
    wi["A5"] = f"Business Type: {business_type.replace('_', ' ').title()}"
    wi["A7"] = "Instructions:"
    wi["A7"].font = Font(bold=True)

    instructions = [
        "1. Each role has its own sheet — fill the correct sheet per employee.",
        "2. employee_id must be UNIQUE across ALL sheets in this file.",
        "3. days_present must not exceed days_assigned.",
        "4. manager_rating must be between 1.0 and 5.0.",
        "5. Do NOT rename or delete column headers.",
        "6. Sheet names must match role names exactly.",
        "7. Leave optional columns as 0 if not applicable.",
        "8. Row 2 is a sample row — overwrite or delete it.",
        "9. Each role sheet has a different sample EMP ID (EMP001, EMP002 etc.).",
        "10. When filling real data, use your actual employee IDs consistently.",
    ]
    for i, line in enumerate(instructions, start=8):
        wi[f"A{i}"]      = line
        wi[f"A{i}"].fill = _INST_FILL

    wi.column_dimensions["A"].width = 80

    # ── role sheets ───────────────────────────────────────────────────────────
    for sheet_index, role in enumerate(get_roles(business_type)):
        sheet_name = role.replace("_", " ").title()
        ws         = wb.create_sheet(sheet_name)
        columns    = get_expected_columns(business_type, role)

        # header row
        for ci, col in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=ci, value=col)
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(18, len(col) + 4)

        # sample data row — unique EMP ID per sheet
        for ci, col in enumerate(columns, start=1):
            cell      = ws.cell(
                row=2, column=ci,
                value=_sample_val(col, role, sheet_index)
            )
            cell.fill = _SAMPLE_FILL

        # freeze header row
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()